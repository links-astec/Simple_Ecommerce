import hmac
import hashlib
import html as html_module
import json
import logging
import requests
from django.conf import settings
from django.core.exceptions import ValidationError as DjangoValidationError
from django.core.validators import validate_email
from django.db import transaction, models
from django.http import HttpResponse, Http404, JsonResponse
from django.utils import timezone
from django.core.mail import send_mail as _django_send_mail
from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_page
from rest_framework import generics, status
from rest_framework.decorators import api_view
from rest_framework.permissions import BasePermission
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser
from .models import Category, Product, ProductImage, Order, OrderItem, SiteSettings
from .serializers import (
    CategorySerializer, ProductListSerializer, ProductDetailSerializer,
    OrderCreateSerializer, OrderSerializer
)

logger = logging.getLogger(__name__)


def send_email(subject, message, recipient_list, from_email=None, fail_silently=False):
    brevo_key = getattr(settings, 'BREVO_API_KEY', '')
    sender_name = getattr(settings, 'EMAIL_SENDER_NAME', "Bel's Haven")
    sender_email = from_email or getattr(settings, 'EMAIL_SENDER_ADDRESS', '') or getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@belshaven.com')
    if '<' in sender_email:
        sender_email = sender_email.split('<')[1].rstrip('>')

    if brevo_key:
        resp = requests.post(
            'https://api.brevo.com/v3/smtp/email',
            headers={'api-key': brevo_key, 'Content-Type': 'application/json'},
            json={
                'sender': {'name': sender_name, 'email': sender_email},
                'to': [{'email': e} for e in recipient_list],
                'subject': subject,
                'textContent': message,
            },
            timeout=15,
        )
        if resp.status_code >= 400:
            err = resp.text
            logger.error('Brevo email failed (%s): %s', resp.status_code, err)
            if not fail_silently:
                raise Exception(f'Email send failed: {err}')
        return
    _django_send_mail(subject=subject, message=message, from_email=from_email or settings.DEFAULT_FROM_EMAIL,
                      recipient_list=recipient_list, fail_silently=fail_silently)


ALLOWED_IMAGE_TYPES = {'image/jpeg', 'image/png', 'image/webp', 'image/gif'}
MAX_IMAGE_SIZE = 10 * 1024 * 1024  # 10 MB


# ── Admin auth ───────────────────────────────────────────────────────────────

class IsAdminKey(BasePermission):
    """Require X-Admin-Key header matching ADMIN_API_KEY setting."""
    def has_permission(self, request, view):
        key = request.headers.get('X-Admin-Key', '')
        expected = getattr(settings, 'ADMIN_API_KEY', '')
        if not expected or not key or key != expected:
            if key:
                logger.warning('Invalid admin key from %s', request.META.get('REMOTE_ADDR'))
            return False
        return True


class AdminLoginView(APIView):
    def post(self, request):
        password = request.data.get('password', '')
        expected = getattr(settings, 'ADMIN_API_KEY', '')
        if not expected:
            return Response({'error': 'Admin not configured'}, status=503)
        if password == expected:
            return Response({'token': expected})
        logger.warning('Failed admin login from %s', request.META.get('REMOTE_ADDR'))
        return Response({'error': 'Invalid password'}, status=401)


# ── Categories ──────────────────────────────────────────────────────────────

class CategoryListView(generics.ListCreateAPIView):
    serializer_class = CategorySerializer

    def get_queryset(self):
        return Category.objects.annotate(
            _product_count=models.Count(
                'products',
                filter=models.Q(products__status='active', products__parent__isnull=True)
            )
        )

    def get_permissions(self):
        if self.request.method not in ('GET', 'HEAD', 'OPTIONS'):
            return [IsAdminKey()]
        return []

    @method_decorator(cache_page(120))
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)


class CategoryDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [IsAdminKey]


# ── Products ─────────────────────────────────────────────────────────────────

class ProductListView(generics.ListCreateAPIView):
    serializer_class = ProductListSerializer

    def get_permissions(self):
        if self.request.method not in ('GET', 'HEAD', 'OPTIONS'):
            return [IsAdminKey()]
        return []

    def get_queryset(self):
        qs = Product.objects.select_related('category', 'parent').prefetch_related('images', 'variants')
        is_admin = IsAdminKey().has_permission(self.request, self)
        if not is_admin:
            qs = qs.filter(status='active', parent__isnull=True)
        else:
            variants_of = self.request.query_params.get('variants_of')
            if variants_of:
                qs = qs.filter(parent__slug=variants_of)
            elif not self.request.query_params.get('include_variants'):
                qs = qs.filter(parent__isnull=True)
        category = self.request.query_params.get('category')
        product_type = self.request.query_params.get('type')
        featured = self.request.query_params.get('featured')
        search = self.request.query_params.get('search')
        if category:
            qs = qs.filter(category__slug=category)
        if product_type:
            qs = qs.filter(product_type=product_type)
        if featured == 'true':
            qs = qs.filter(is_featured=True)
        if search:
            qs = qs.filter(name__icontains=search)
        return qs

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx['request'] = self.request
        return ctx

    def get(self, request, *args, **kwargs):
        if not IsAdminKey().has_permission(request, self):
            return cache_page(120)(super().get)(request, *args, **kwargs)
        return super().get(request, *args, **kwargs)

    def create(self, request, *args, **kwargs):
        from django.utils.text import slugify as django_slugify
        data = request.data.copy()
        images = data.pop('images', [])
        variants_data = data.pop('variants', [])
        if isinstance(images, str):
            import json as _json
            try:
                images = _json.loads(images)
            except Exception:
                images = []
        if isinstance(variants_data, str):
            import json as _json
            try:
                variants_data = _json.loads(variants_data)
            except Exception:
                variants_data = []
        serializer = ProductListSerializer(data=data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        product = serializer.save()
        if images:
            product.images.set(images)
        for v in variants_data:
            label = v.get('variant_label', '')
            v_slug = f"{product.slug}-{django_slugify(label)}" if label else f"{product.slug}-v"
            if Product.objects.filter(slug=v_slug).exists():
                import secrets
                v_slug = f"{v_slug}-{secrets.token_hex(2)}"
            variant = Product.objects.create(
                parent=product,
                name=product.name,
                slug=v_slug,
                description=product.description,
                category=product.category,
                product_type=v.get('product_type', product.product_type),
                status='active',
                variant_label=label,
                price=v.get('price', product.price),
                stock_quantity=v.get('stock_quantity', 0),
                shipping_fee=v.get('shipping_fee', product.shipping_fee),
                delivery_timeframe=v.get('delivery_timeframe', product.delivery_timeframe),
                preorder_eta=v.get('preorder_eta', product.preorder_eta),
                preorder_shipping_fee=v.get('preorder_shipping_fee', product.preorder_shipping_fee),
            )
            v_images = v.get('images', [])
            if v_images:
                variant.images.set(v_images)
        return Response(
            ProductDetailSerializer(product, context={'request': request}).data,
            status=status.HTTP_201_CREATED
        )


class ProductDetailView(generics.RetrieveUpdateDestroyAPIView):
    lookup_field = 'slug'

    def get_permissions(self):
        if self.request.method not in ('GET', 'HEAD', 'OPTIONS'):
            return [IsAdminKey()]
        return []

    def get_queryset(self):
        if self.request.method == 'GET':
            return Product.objects.filter(status='active').select_related('category', 'parent').prefetch_related('images', 'variants', 'variants__images')
        return Product.objects.all().select_related('category', 'parent').prefetch_related('images', 'variants', 'variants__images')

    def get_serializer_class(self):
        if self.request.method == 'GET':
            return ProductDetailSerializer
        return ProductListSerializer

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx['request'] = self.request
        return ctx

    def update(self, request, *args, **kwargs):
        from django.utils.text import slugify as django_slugify
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        data = request.data.copy()
        images = data.pop('images', None)
        variants_data = data.pop('variants', None)
        if isinstance(images, str):
            import json as _json
            try:
                images = _json.loads(images)
            except Exception:
                images = None
        if isinstance(variants_data, str):
            import json as _json
            try:
                variants_data = _json.loads(variants_data)
            except Exception:
                variants_data = None
        serializer = ProductListSerializer(instance, data=data, partial=partial, context={'request': request})
        serializer.is_valid(raise_exception=True)
        product = serializer.save()
        if images is not None:
            product.images.set(images)
        if variants_data is not None:
            existing_ids = set()
            for v in variants_data:
                v_id = v.get('id')
                if v_id:
                    try:
                        variant = Product.objects.get(id=v_id, parent=product)
                        for field in ('variant_label', 'price', 'stock_quantity', 'shipping_fee',
                                      'delivery_timeframe', 'preorder_eta', 'preorder_shipping_fee', 'product_type'):
                            if field in v:
                                setattr(variant, field, v[field])
                        variant.name = product.name
                        variant.save()
                        existing_ids.add(str(variant.id))
                    except Product.DoesNotExist:
                        v_id = None
                if not v_id:
                    label = v.get('variant_label', '')
                    v_slug = f"{product.slug}-{django_slugify(label)}" if label else f"{product.slug}-v"
                    if Product.objects.filter(slug=v_slug).exists():
                        import secrets
                        v_slug = f"{v_slug}-{secrets.token_hex(2)}"
                    variant = Product.objects.create(
                        parent=product, name=product.name, slug=v_slug,
                        description=product.description, category=product.category,
                        product_type=v.get('product_type', product.product_type),
                        status='active', variant_label=label,
                        price=v.get('price', product.price),
                        stock_quantity=v.get('stock_quantity', 0),
                        shipping_fee=v.get('shipping_fee', product.shipping_fee),
                        delivery_timeframe=v.get('delivery_timeframe', product.delivery_timeframe),
                        preorder_eta=v.get('preorder_eta', product.preorder_eta),
                        preorder_shipping_fee=v.get('preorder_shipping_fee', product.preorder_shipping_fee),
                    )
                    existing_ids.add(str(variant.id))
                v_images = v.get('images')
                if v_images is not None:
                    variant.images.set(v_images)
            product.variants.exclude(id__in=existing_ids).delete()
        return Response(ProductDetailSerializer(product, context={'request': request}).data)


# ── Product Images ────────────────────────────────────────────────────────────

class ProductImageUploadView(APIView):
    parser_classes = [MultiPartParser, FormParser]
    permission_classes = [IsAdminKey]

    def post(self, request):
        image_file = request.FILES.get('image')
        if not image_file:
            return Response({'error': 'No image provided'}, status=400)
        if image_file.content_type not in ALLOWED_IMAGE_TYPES:
            return Response({'error': 'Invalid file type. Allowed: JPEG, PNG, WebP, GIF'}, status=400)
        if image_file.size > MAX_IMAGE_SIZE:
            return Response({'error': 'File too large. Maximum 10 MB allowed'}, status=400)
        is_primary = request.data.get('is_primary', 'false') == 'true'
        img = ProductImage.objects.create(image=image_file, is_primary=is_primary)
        url = img.image.url
        if not url.startswith('http'):
            url = request.build_absolute_uri(url)
        return Response({'id': img.id, 'url': url, 'is_primary': img.is_primary}, status=201)


# ── Orders ────────────────────────────────────────────────────────────────────

class OrderCreateView(APIView):
    def post(self, request):
        serializer = OrderCreateSerializer(data=request.data)
        if serializer.is_valid():
            order = serializer.save()
            return Response({
                'order_id': str(order.id),
                'reference': order.reference,
                'total_amount': float(order.total_amount),
                'email': order.customer_email,
            }, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class AdminOrderListView(generics.ListAPIView):
    serializer_class = OrderSerializer
    permission_classes = [IsAdminKey]

    def get_queryset(self):
        return Order.objects.prefetch_related('items').order_by('-created_at')


class OrderDetailView(APIView):
    def get_permissions(self):
        # GET is public (user tracks own order); PATCH is admin-only
        if self.request.method == 'PATCH':
            return [IsAdminKey()]
        return []

    def get(self, request, reference):
        try:
            order = Order.objects.prefetch_related('items').get(reference=reference)
            return Response(OrderSerializer(order).data)
        except Order.DoesNotExist:
            return Response({'error': 'Order not found'}, status=404)

    def patch(self, request, reference):
        try:
            order = Order.objects.get(reference=reference)
            allowed = {'status', 'notes'}
            for field in allowed:
                if field in request.data:
                    setattr(order, field, request.data[field])
            order.save()
            logger.info('Admin updated order %s: status=%s', reference, order.status)
            return Response(OrderSerializer(order).data)
        except Order.DoesNotExist:
            return Response({'error': 'Order not found'}, status=404)


_email_codes = {}

class OrdersByEmailView(APIView):
    def options(self, request, *args, **kwargs):
        resp = HttpResponse(status=200)
        origin = request.META.get('HTTP_ORIGIN', '*')
        resp['Access-Control-Allow-Origin'] = origin
        resp['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        resp['Access-Control-Allow-Headers'] = 'Content-Type, X-Admin-Key, Cache-Control, Pragma'
        resp['Access-Control-Max-Age'] = '86400'
        return resp

    def post(self, request):
        action = request.data.get('action', '')
        email = request.data.get('email', '').strip().lower()

        if not email:
            return Response({'error': 'Email is required.'}, status=400)
        try:
            validate_email(email)
        except DjangoValidationError:
            return Response({'error': 'Invalid email address.'}, status=400)

        if action == 'send_code':
            if not Order.objects.filter(customer_email__iexact=email).exists():
                return Response({'error': 'No orders found for this email.'}, status=404)

            import secrets as _secrets
            code = _secrets.token_hex(3).upper()
            _email_codes[email] = {'code': code, 'ts': timezone.now()}

            for k in list(_email_codes):
                if (timezone.now() - _email_codes[k]['ts']).total_seconds() > 600:
                    del _email_codes[k]

            try:
                send_email(
                    subject="Your verification code - Bel's Haven",
                    message="Your order lookup code is: %s\n\nThis code expires in 10 minutes.\n\n- Bel's Haven" % code,
                    recipient_list=[email],
                )
                return Response({'status': 'code_sent'})
            except Exception as e:
                logger.exception('Email send failed: %s', e)
                return Response({'error': 'Could not send verification email. Please try again or use your order reference number instead.'}, status=500)

        if action == 'verify':
            code = request.data.get('code', '').strip().upper()
            if not code:
                return Response({'error': 'Verification code is required.'}, status=400)
            entry = _email_codes.get(email)
            if not entry or entry['code'] != code:
                return Response({'error': 'Invalid code.'}, status=403)
            if (timezone.now() - entry['ts']).total_seconds() > 600:
                del _email_codes[email]
                return Response({'error': 'Code expired. Request a new one.'}, status=403)
            del _email_codes[email]
            orders = Order.objects.filter(
                customer_email__iexact=email
            ).prefetch_related('items').order_by('-created_at')
            return Response(OrderSerializer(orders, many=True).data)

        return Response({'error': 'Invalid action. Use send_code or verify.'}, status=400)


# ── Payment ───────────────────────────────────────────────────────────────────

class PaystackInitializeView(APIView):
    def post(self, request):
        order_id = request.data.get('order_id')
        try:
            order = Order.objects.get(id=order_id)
        except Order.DoesNotExist:
            return Response({'error': 'Order not found'}, status=404)
        if order.payment_verified:
            return Response({'error': 'Already paid'}, status=400)

        amount_kobo = int(float(order.total_amount) * 100)
        base_callback = request.data.get('callback_url') or f"{settings.FRONTEND_URL}/payment/verify"
        callback_url = f"{base_callback}?reference={order.reference}"
        payload = {
            'email': order.customer_email,
            'amount': amount_kobo,
            'reference': order.reference,
            'callback_url': callback_url,
            'metadata': {'order_id': str(order.id), 'customer_name': order.customer_name}
        }
        headers = {'Authorization': f'Bearer {settings.PAYSTACK_SECRET_KEY}', 'Content-Type': 'application/json'}
        try:
            resp = requests.post('https://api.paystack.co/transaction/initialize', json=payload, headers=headers, timeout=10)
            data = resp.json()
            if data.get('status'):
                return Response({
                    'authorization_url': data['data']['authorization_url'],
                    'access_code': data['data']['access_code'],
                    'reference': data['data']['reference'],
                })
            return Response({'error': data.get('message', 'Payment initialization failed')}, status=400)
        except Exception:
            logger.exception('Payment initialization error for order %s', order_id)
            return Response({'error': 'Payment initialization failed. Please try again.'}, status=500)


class PaystackVerifyView(APIView):
    def get(self, request):
        reference = request.query_params.get('reference')
        if not reference:
            return Response({'error': 'Reference required'}, status=400)
        try:
            order = Order.objects.get(reference=reference)
        except Order.DoesNotExist:
            return Response({'error': 'Order not found'}, status=404)
        if order.payment_verified:
            return Response({'status': 'already_verified', 'order': OrderSerializer(order).data})

        headers = {'Authorization': f'Bearer {settings.PAYSTACK_SECRET_KEY}'}
        try:
            resp = requests.get(f'https://api.paystack.co/transaction/verify/{reference}', headers=headers, timeout=10)
            data = resp.json()
            if data.get('status') and data['data']['status'] == 'success':
                with transaction.atomic():
                    order = Order.objects.select_for_update().get(reference=reference)
                    if order.payment_verified:
                        return Response({'status': 'already_verified', 'order': OrderSerializer(order).data})
                    order.payment_verified = True
                    order.status = Order.STATUS_PAID
                    order.payment_date = timezone.now()
                    order.paystack_reference = reference
                    order.save()
                    for item in order.items.select_related('product'):
                        if item.product and item.product_type == 'available':
                            Product.objects.filter(pk=item.product.pk).update(
                                stock_quantity=models.F('stock_quantity') - item.quantity
                            )
                _send_customer_receipt(order)
                _send_admin_notification(order)
                return Response({'status': 'success', 'order': OrderSerializer(order).data})
            return Response({'status': 'failed', 'message': 'Payment not successful'}, status=400)
        except Exception:
            logger.exception('Payment verification error for reference %s', reference)
            return Response({'error': 'Verification failed. Please contact support.'}, status=500)


class PaystackWebhookView(APIView):
    def post(self, request):
        payload = request.body
        sig = request.headers.get('x-paystack-signature', '')
        expected = hmac.new(settings.PAYSTACK_SECRET_KEY.encode('utf-8'), payload, hashlib.sha512).hexdigest()
        if sig != expected:
            return Response(status=400)
        event = json.loads(payload)
        if event.get('event') == 'charge.success':
            ref = event['data']['reference']
            try:
                order = Order.objects.get(reference=ref, payment_verified=False)
                order.payment_verified = True
                order.status = Order.STATUS_PAID
                order.payment_date = timezone.now()
                order.save()
                _send_customer_receipt(order)
                _send_admin_notification(order)
            except Order.DoesNotExist:
                pass
        return Response(status=200)


# ── Stats ─────────────────────────────────────────────────────────────────────

@api_view(['GET'])
def store_stats(request):
    return Response({
        'total_products': Product.objects.filter(status='active', parent__isnull=True).count(),
        'available_products': Product.objects.filter(status='active', product_type='available', parent__isnull=True).count(),
        'preorder_products': Product.objects.filter(status='active', product_type='preorder', parent__isnull=True).count(),
        'categories': Category.objects.count(),
    })


@api_view(['GET'])
def email_check(request):
    if not IsAdminKey().has_permission(request, None):
        return Response({'error': 'Unauthorized'}, status=403)
    host_user = getattr(settings, 'EMAIL_HOST_USER', '')
    host = getattr(settings, 'EMAIL_HOST', '')
    port = getattr(settings, 'EMAIL_PORT', '')
    from_email = getattr(settings, 'DEFAULT_FROM_EMAIL', '')
    has_password = bool(getattr(settings, 'EMAIL_HOST_PASSWORD', ''))
    return Response({
        'EMAIL_HOST': host,
        'EMAIL_PORT': port,
        'EMAIL_HOST_USER': host_user,
        'EMAIL_HOST_PASSWORD': '***set***' if has_password else '***NOT SET***',
        'DEFAULT_FROM_EMAIL': from_email,
    })


# ── Email Helpers ─────────────────────────────────────────────────────────────

def _send_customer_receipt(order):
    whatsapp = settings.WHATSAPP_NUMBER.replace('+', '').replace(' ', '')
    items_text = '\n'.join([f"  - {i.product_name} x{i.quantity} @ GH₵{i.unit_price:,.2f}" for i in order.items.all()])
    wa_msg = (
        f"Hello! I just placed an order on Bel's Haven.\n"
        f"Order Reference: *{order.reference}*\n"
        f"Name: {order.customer_name}\n"
        f"Amount Paid: GH₵{order.total_amount:,.2f}\n"
        f"Please confirm my order. Thank you! 🛍️"
    )
    wa_url = f"https://wa.me/{whatsapp}?text={requests.utils.quote(wa_msg)}"
    body = f"""Hello {order.customer_name},

Thank you for shopping at Bel's Haven! 🖤✨
Your order has been received and payment confirmed.

Reference:  {order.reference}
Date:       {order.created_at.strftime('%B %d, %Y at %I:%M %p')}

Items:
{items_text}

Shipping:   GH₵{order.shipping_fee:,.2f}
Total Paid: GH₵{order.total_amount:,.2f}

Delivery To:
{order.delivery_address}, {order.city}, {order.state}

━━━━━━━━━━━━━━━━━━━━━━
Validate your order on WhatsApp: {wa_url}

With love, Bel's Haven 🌿
"""
    try:
        send_email(subject=f"Order Confirmed - {order.reference} | Bel's Haven", message=body,
                   recipient_list=[order.customer_email], fail_silently=True)
    except Exception:
        pass


def health(request):
    return JsonResponse({"status": "ok"})


def product_share(request, slug):
    """Serve an OG-tagged HTML page for WhatsApp/social link previews, then redirect to the SPA."""
    try:
        product = Product.objects.prefetch_related('images').get(slug=slug, status='active')
    except Product.DoesNotExist:
        raise Http404

    primary_img = product.images.filter(is_primary=True).first() or product.images.first()
    image_url = ''
    if primary_img:
        image_url = primary_img.image.url
        if not image_url.startswith('http'):
            image_url = request.build_absolute_uri(image_url)

    frontend_url = getattr(settings, 'FRONTEND_URL', '').split(',')[0].strip().rstrip('/')
    product_url = f"{frontend_url}/shop/{product.slug}" if frontend_url else request.build_absolute_uri(f'/shop/{product.slug}')

    if product.parent:
        product_url = f"{frontend_url}/shop/{product.parent.slug}" if frontend_url else request.build_absolute_uri(f'/shop/{product.parent.slug}')

    esc = html_module.escape
    display_name = product.name
    if product.variant_label:
        display_name = f"{product.name} ({product.variant_label})"
    name = esc(display_name)
    description = esc((product.description or '')[:200])
    price_display = f"GH₵{float(product.price):,.2f}"

    og_image = f'<meta property="og:image" content="{esc(image_url)}">\n  <meta property="og:image:width" content="1200">\n  <meta property="og:image:height" content="630">\n  <meta name="twitter:image" content="{esc(image_url)}">' if image_url else ''

    page = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>{name} — Bel&apos;s Haven</title>
  <meta name="description" content="{description}">
  <meta property="og:type" content="product">
  <meta property="og:title" content="{name} — Bel&apos;s Haven">
  <meta property="og:description" content="{description}">
  <meta property="og:url" content="{esc(product_url)}">
  <meta property="og:site_name" content="Bel&apos;s Haven">
  {og_image}
  <meta property="product:price:amount" content="{float(product.price)}">
  <meta property="product:price:currency" content="GHS">
  <meta name="twitter:card" content="summary_large_image">
  <meta name="twitter:title" content="{name} — Bel&apos;s Haven">
  <meta name="twitter:description" content="{description}">
  <meta http-equiv="refresh" content="0;url={esc(product_url)}">
  <link rel="canonical" href="{esc(product_url)}">
  <style>body{{font-family:sans-serif;display:flex;align-items:center;justify-content:center;height:100vh;margin:0;background:#faf8f4;color:#4a3f2f}}</style>
</head>
<body>
  <script>window.location.replace({json.dumps(product_url)});</script>
  <p>{name} — {price_display} &mdash; <a href="{esc(product_url)}">View product</a></p>
</body>
</html>"""

    return HttpResponse(page, content_type='text/html; charset=utf-8')


class DataExportView(APIView):
    def get(self, request):
        key = request.headers.get('X-Admin-Key', '') or request.query_params.get('key', '')
        expected = getattr(settings, 'ADMIN_API_KEY', '')
        if not expected or not key or key != expected:
            return Response({'error': 'Unauthorized'}, status=403)

        fmt = request.query_params.get('format', 'json')
        now = timezone.now()
        date_str = now.strftime('%Y-%m-%d')

        categories = list(Category.objects.values('id', 'name', 'slug', 'description'))

        products = []
        for p in Product.objects.prefetch_related('images', 'category').all():
            img_urls = [img.image.url for img in p.images.all()]
            products.append({
                'id': str(p.id), 'name': p.name, 'slug': p.slug,
                'description': p.description, 'price': str(p.price),
                'shipping_fee': str(p.shipping_fee), 'product_type': p.product_type,
                'status': p.status, 'stock_quantity': p.stock_quantity,
                'delivery_timeframe': p.delivery_timeframe,
                'preorder_eta': p.preorder_eta,
                'preorder_shipping_fee': str(p.preorder_shipping_fee),
                'is_featured': p.is_featured,
                'category': p.category.name if p.category else '',
                'variant_label': p.variant_label,
                'parent_slug': p.parent.slug if p.parent else '',
                'images': img_urls,
                'created_at': p.created_at.isoformat() if p.created_at else '',
            })

        orders = []
        order_items_flat = []
        customers_seen = set()
        customers = []
        for o in Order.objects.prefetch_related('items').all():
            orders.append({
                'id': str(o.id), 'reference': o.reference,
                'customer_name': o.customer_name, 'customer_email': o.customer_email,
                'customer_phone': o.customer_phone,
                'delivery_address': o.delivery_address, 'city': o.city,
                'state': o.state, 'country': o.country,
                'status': o.status, 'total_amount': str(o.total_amount),
                'shipping_fee': str(o.shipping_fee),
                'payment_verified': o.payment_verified,
                'payment_date': o.payment_date.isoformat() if o.payment_date else '',
                'notes': o.notes,
                'created_at': o.created_at.isoformat() if o.created_at else '',
            })
            for item in o.items.all():
                order_items_flat.append({
                    'order_reference': o.reference,
                    'product_name': item.product_name,
                    'product_type': item.product_type,
                    'quantity': item.quantity,
                    'unit_price': str(item.unit_price),
                    'shipping_fee': str(item.shipping_fee),
                })
            email_key = o.customer_email.lower()
            if email_key not in customers_seen:
                customers_seen.add(email_key)
                customers.append({
                    'name': o.customer_name, 'email': o.customer_email,
                    'phone': o.customer_phone,
                    'address': o.delivery_address, 'city': o.city,
                    'state': o.state, 'country': o.country,
                })

        if fmt == 'excel':
            try:
                import openpyxl
                from io import BytesIO
                wb = openpyxl.Workbook()

                def write_sheet(ws, headers, rows):
                    ws.append(headers)
                    for r in rows:
                        ws.append([str(r.get(h, '') or '') for h in headers])

                ws_cat = wb.active
                ws_cat.title = 'Categories'
                write_sheet(ws_cat, ['id', 'name', 'slug', 'description'], categories)

                ws_prod = wb.create_sheet('Products')
                prod_headers = ['id', 'name', 'slug', 'category', 'price', 'shipping_fee',
                               'product_type', 'status', 'stock_quantity', 'variant_label',
                               'parent_slug', 'delivery_timeframe', 'preorder_eta',
                               'is_featured', 'images', 'created_at']
                ws_prod.append(prod_headers)
                for p in products:
                    row = [str(p.get(h, '') or '') for h in prod_headers]
                    img_idx = prod_headers.index('images')
                    row[img_idx] = ', '.join(p.get('images', []))
                    ws_prod.append(row)

                ws_ord = wb.create_sheet('Orders')
                ord_headers = ['reference', 'customer_name', 'customer_email', 'customer_phone',
                              'delivery_address', 'city', 'state', 'country', 'status',
                              'total_amount', 'shipping_fee', 'payment_verified',
                              'payment_date', 'notes', 'created_at']
                write_sheet(ws_ord, ord_headers, orders)

                ws_items = wb.create_sheet('Order Items')
                item_headers = ['order_reference', 'product_name', 'product_type',
                               'quantity', 'unit_price', 'shipping_fee']
                write_sheet(ws_items, item_headers, order_items_flat)

                ws_cust = wb.create_sheet('Customers')
                cust_headers = ['name', 'email', 'phone', 'address', 'city', 'state', 'country']
                write_sheet(ws_cust, cust_headers, customers)

                buf = BytesIO()
                wb.save(buf)
                buf.seek(0)
                response = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="bels-haven-backup-{date_str}.xlsx"'
                return response
            except Exception as e:
                logger.exception('Excel export failed')
                return Response({'error': f'Excel export failed: {str(e)}'}, status=500)

        if fmt == 'pdf':
            try:
                from io import BytesIO
                from reportlab.lib import colors
                from reportlab.lib.pagesizes import A4, landscape
                from reportlab.lib.units import mm
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer
                from reportlab.platypus import Paragraph as RLParagraph
                from reportlab.lib.styles import getSampleStyleSheet

                def trunc(val, mx=80):
                    s = str(val) if val is not None else ''
                    return s[:mx] + '...' if len(s) > mx else s

                buf = BytesIO()
                doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=12*mm, rightMargin=12*mm, topMargin=15*mm, bottomMargin=12*mm)
                styles = getSampleStyleSheet()
                elements = []

                elements.append(RLParagraph("Bel's Haven - Data Backup".replace("'", "&#8217;"), styles['Title']))
                elements.append(RLParagraph(f"Exported: {now.strftime('%B %d, %Y at %H:%M')}", styles['Normal']))
                elements.append(Spacer(1, 8*mm))

                header_style = TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f4f0e8')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#4a3f2f')),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('FONTSIZE', (0, 1), (-1, -1), 7),
                    ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#e8e0d0')),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('TOPPADDING', (0, 0), (-1, -1), 3),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#faf8f4')]),
                ])

                def add_section(title, headers, rows):
                    elements.append(RLParagraph(title.replace("'", "&#8217;").replace("&", "&amp;").replace("<", "&lt;"), styles['Heading2']))
                    if not rows:
                        elements.append(RLParagraph("No data", styles['Normal']))
                        elements.append(Spacer(1, 5*mm))
                        return
                    table_data = [headers]
                    for r in rows:
                        table_data.append([trunc(r.get(h, '')) for h in headers])
                    t = Table(table_data, repeatRows=1)
                    t.setStyle(header_style)
                    elements.append(t)
                    elements.append(Spacer(1, 6*mm))

                add_section(f"Categories ({len(categories)})",
                            ['name', 'slug', 'description'], categories)

                parent_products = [p for p in products if not p['parent_slug']]
                add_section(f"Products ({len(parent_products)})",
                            ['name', 'category', 'price', 'stock_quantity', 'product_type', 'status'],
                            parent_products)

                variant_products = [p for p in products if p['parent_slug']]
                if variant_products:
                    add_section(f"Product Variants ({len(variant_products)})",
                                ['parent_slug', 'variant_label', 'price', 'stock_quantity', 'shipping_fee'],
                                variant_products)

                add_section(f"Orders ({len(orders)})",
                            ['reference', 'customer_name', 'customer_email', 'status', 'total_amount', 'created_at'],
                            orders)

                add_section(f"Customers ({len(customers)})",
                            ['name', 'email', 'phone', 'city', 'country'],
                            customers)

                doc.build(elements)
                buf.seek(0)
                response = HttpResponse(buf.getvalue(), content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="bels-haven-backup-{date_str}.pdf"'
                return response
            except Exception as e:
                logger.exception('PDF export failed')
                return Response({'error': f'PDF generation failed: {str(e)}'}, status=500)

        data = {
            'exported_at': now.isoformat(),
            'categories': categories,
            'products': products,
            'orders': orders,
            'order_items': order_items_flat,
            'customers': customers,
        }
        response = HttpResponse(
            json.dumps(data, indent=2, default=str),
            content_type='application/json'
        )
        response['Content-Disposition'] = f'attachment; filename="bels-haven-backup-{date_str}.json"'
        return response


class SiteSettingsView(APIView):
    def get(self, request):
        s = SiteSettings.get()
        return Response({'maintenance': s.maintenance_mode})

    def patch(self, request):
        if not IsAdminKey().has_permission(request, self):
            return Response({'error': 'Unauthorized'}, status=403)
        s = SiteSettings.get()
        if 'maintenance' in request.data:
            s.maintenance_mode = bool(request.data['maintenance'])
            s.save()
        return Response({'maintenance': s.maintenance_mode})


class SendCustomerMessageView(APIView):
    permission_classes = [IsAdminKey]

    def post(self, request):
        email = request.data.get('email', '').strip()
        subject = request.data.get('subject', "Message from Bel's Haven").strip()
        message = request.data.get('message', '').strip()
        if not email or not message:
            return Response({'error': 'Email and message required'}, status=400)
        try:
            validate_email(email)
        except DjangoValidationError:
            return Response({'error': 'Invalid email address'}, status=400)
        try:
            send_email(
                subject=subject,
                message=f"{message}\n\n- Bel's Haven",
                recipient_list=[email],
            )
            logger.info('Admin sent message to %s', email)
            return Response({'status': 'sent'})
        except Exception:
            logger.exception('Failed to send message to %s', email)
            return Response({'error': 'Failed to send message. Please try again.'}, status=500)

def _send_admin_notification(order):
    notify_email = getattr(settings, 'NOTIFY_EMAIL', None) or getattr(settings, 'ADMIN_EMAIL', None)
    if not notify_email:
        return
    items_text = '\n'.join([f"  • {i.product_name} (x{i.quantity}) – GH₵{i.unit_price:,.2f} [{i.product_type}]" for i in order.items.all()])
    body = f"""New order on Bel's Haven!

{order.customer_name} | {order.customer_email} | {order.customer_phone}
{order.delivery_address}, {order.city}, {order.state}

Order #{order.reference}
{items_text}
Shipping: GH₵{order.shipping_fee:,.2f}
TOTAL: GH₵{order.total_amount:,.2f}
Notes: {order.notes or 'None'}

Manage at: {settings.FRONTEND_URL}/manage
"""
    try:
        send_email(subject=f"New Order: {order.reference} - GH₵{order.total_amount:,.2f}", message=body,
                   recipient_list=[notify_email], fail_silently=True)
    except Exception:
        pass
